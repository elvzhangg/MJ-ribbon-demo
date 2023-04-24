import json
import openpyxl
import uuid
import argparse


def read_xlsx(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    data = []
    headers = []

    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = row

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for idx, cell in enumerate(row):
            header = headers[idx]
            if header in ["msg_content", "index", "type", "img_url"]:
                record[header] = cell
        record['id'] = str(uuid.uuid4())  # Add a unique id to each record
        data.append(record)

    return data


# Parse command line arguments
parser = argparse.ArgumentParser(description='Process XLSX file to JSON.')
parser.add_argument('input_filename', type=str, help='Input XLSX file name')
parser.add_argument('output_filename', type=str, help='Output JSON file name')

args = parser.parse_args()

xlsx_filename = args.input_filename
json_filename = args.output_filename

data = read_xlsx(xlsx_filename)

with open(json_filename, 'w') as json_file:
    json.dump(data, json_file, indent=2)

print(f"Data extracted from {xlsx_filename} and saved to {json_filename}.")
